"""
parser/bank_parser.py

Config-driven Excel parser for bank statements.
Reads column mappings from a YAML file (e.g. banks/santander.yaml).
Zero hardcoded column logic here — all structural knowledge lives in the YAML.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import yaml


# ---------------------------------------------------------------------------
# Domain model — what comes out of the parser
# ---------------------------------------------------------------------------

@dataclass
class BankMetadata:
    bank_id: str
    bank_name: str
    account_number: str | None
    account_holder: str | None
    current_balance: float | None
    export_date: str | None
    currency: str


@dataclass
class RawTransaction:
    date_operation: date
    date_value: date
    description: str
    amount: float
    balance: float
    currency: str
    transaction_type: str       # detected from YAML patterns
    is_reversal: bool
    row_index: int              # original spreadsheet row (1-based), useful for debugging


@dataclass
class ParseResult:
    metadata: BankMetadata
    transactions: list[RawTransaction]
    parse_warnings: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------

class BankParser:
    """
    Reads a bank statement Excel file using column mappings from a YAML config.
    Supports any bank that exports to xlsx by providing its own YAML.
    """

    def __init__(self, config_path: str | Path):
        self._cfg = self._load_config(Path(config_path))

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def parse(self, file_path: str | Path) -> ParseResult:
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        sheet_cfg = self._cfg["sheet"]
        ws = wb.worksheets[sheet_cfg["index"]]

        metadata = self._extract_metadata(ws)
        transactions, warnings = self._extract_transactions(ws)

        wb.close()
        return ParseResult(
            metadata=metadata,
            transactions=transactions,
            parse_warnings=warnings,
        )

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _load_config(path: Path) -> dict:
        with path.open(encoding="utf-8") as f:
            return yaml.safe_load(f)

    def _extract_metadata(self, ws) -> BankMetadata:
        cfg = self._cfg
        meta_cfg = cfg["sheet"]["metadata"]

        def cell(row: int, col: int) -> Any:
            return ws.cell(row=row, column=col).value

        raw_balance = cell(meta_cfg["balance_value_row"], meta_cfg["balance_col"])
        parsed_balance = self._parse_amount(str(raw_balance)) if raw_balance else None

        return BankMetadata(
            bank_id=cfg["bank"]["id"],
            bank_name=cfg["bank"]["name"],
            account_number=str(cell(meta_cfg["account_value_row"], meta_cfg["account_col"]) or "").strip() or None,
            account_holder=str(cell(meta_cfg["holder_value_row"], meta_cfg["holder_col"]) or "").strip() or None,
            current_balance=parsed_balance,
            export_date=str(cell(meta_cfg["export_date_value_row"], meta_cfg["export_date_col"]) or "").strip() or None,
            currency=cfg["bank"]["currency"],
        )

    def _extract_transactions(self, ws) -> tuple[list[RawTransaction], list[str]]:
        cfg = self._cfg
        sheet_cfg = cfg["sheet"]
        col_cfg = cfg["columns"]
        parse_cfg = cfg["parsing"]
        type_patterns = [
            (t["id"], re.compile(t["pattern"], re.IGNORECASE))
            for t in cfg.get("transaction_types", [])
        ]

        data_start = sheet_cfg["data_start_row"]
        date_fmt = parse_cfg["date_format"]
        reversal_prefix = parse_cfg.get("reversal_prefix", "ANULACION").upper()
        min_cols = parse_cfg.get("min_valid_columns", 4)

        transactions: list[RawTransaction] = []
        warnings: list[str] = []

        for row_num, row in enumerate(ws.iter_rows(min_row=data_start, values_only=True), start=data_start):
            # Skip rows that don't have enough populated columns
            non_null = sum(1 for v in row if v is not None)
            if non_null < min_cols:
                continue

            try:
                raw_date_op = str(row[col_cfg["date_operation"] - 1] or "").strip()
                raw_date_val = str(row[col_cfg["date_value"] - 1] or "").strip()
                description = str(row[col_cfg["description"] - 1] or "").strip()
                raw_amount = str(row[col_cfg["amount"] - 1] or "").strip()
                raw_balance = str(row[col_cfg["balance"] - 1] or "").strip()
                currency = str(row[col_cfg["currency"] - 1] or cfg["bank"]["currency"]).strip()

                if not raw_date_op or not description or not raw_amount:
                    continue

                date_op = datetime.strptime(raw_date_op, date_fmt).date()
                date_val = datetime.strptime(raw_date_val, date_fmt).date() if raw_date_val else date_op
                amount = self._parse_amount(raw_amount)
                balance = self._parse_amount(raw_balance) if raw_balance else 0.0

                tx_type = self._detect_type(description, type_patterns)
                is_reversal = description.upper().startswith(reversal_prefix)

                transactions.append(RawTransaction(
                    date_operation=date_op,
                    date_value=date_val,
                    description=description,
                    amount=amount,
                    balance=balance,
                    currency=currency,
                    transaction_type=tx_type,
                    is_reversal=is_reversal,
                    row_index=row_num,
                ))
            except Exception as exc:  # noqa: BLE001
                warnings.append(f"Row {row_num}: skipped — {exc}")

        return transactions, warnings

    def _parse_amount(self, raw: str) -> float:
        """Convert Spanish-locale amount string to float.
        '-14,50€' -> -14.50
        '3.411,06€' -> 3411.06
        """
        parse_cfg = self._cfg["parsing"]
        strip_pat = parse_cfg["amount_strip_pattern"]
        dec_sep = parse_cfg["amount_decimal_separator"]

        is_negative = raw.strip().startswith("-")
        cleaned = re.sub(strip_pat, "", raw.replace("-", ""))
        # Also strip any remaining non-numeric, non-comma chars (e.g. "EUR" in metadata balance)
        cleaned = re.sub(r"[^0-9,]", "", cleaned)
        cleaned = cleaned.replace(dec_sep, ".")
        value = float(cleaned)
        return -value if is_negative else value

    @staticmethod
    def _detect_type(description: str, patterns: list[tuple[str, re.Pattern]]) -> str:
        for type_id, pattern in patterns:
            if pattern.search(description):
                return type_id
        return "unknown"
